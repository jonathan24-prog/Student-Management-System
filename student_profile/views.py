from django.shortcuts import render,redirect,get_object_or_404
from .forms import StudentForm, EnrollForm, SubjectLoadForm,CurriculumForm,UploadExcelForm,SubjectForm,IDsNameForm,StudentFormUpdate, AnnouncementForm
from .mixins import AjaxFormMixin
from django.contrib.auth.decorators import login_required
from django.views.generic import FormView, View
from django.shortcuts import render,redirect,HttpResponse,get_object_or_404
from django.http import JsonResponse
from rest_framework.decorators import api_view,authentication_classes,permission_classes
from rest_framework.response import Response
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated
from accounts.models import CustomUser
from .models  import ActiveSem, DaySched, TimeSched, Room, InstructorLoadSubject, AOSF, Instructor, EnrollbyStudent, Student,Major, Academic_year, Year_level,Course, SubjectsLoaded,Subject,Curriculum,IDandFullname, Student, Announcement, StudentProfile, Subject
from django.conf import settings
from .serializers import UserSerializer, ActiveSemSerializer, InstructorLoadSubjectWithObjectSerializer, RoomSerializer, DaySchedSerializer, TimeSchedSerializer, InstructorLoadSubjectSerializer, AOSFSerializer, StudentsEnrollSerializer, CurriculumSerializer,CurriculumsSerializer, IDandFullnameSerializer, EnrollStudSerlizer, SubjectSerializer,InstructorSerializer, StudentEnrollSerializer, SubjectsLoadedSerializer, CourseSerializer, Academic_yearSerializer, MajorSerializer,Year_levelSerializer,StudentSerializer,SubjectEnrollSerializer,AnnouncementSerializer, StudentProfileSerializer
from rest_framework import generics
# import numpy as np
# import pandas as pd
from django.utils import timezone
from datetime import datetime
from django.db.models import Q
from rest_framework import viewsets
from rest_framework import filters
import os
from os.path import expanduser
from django.conf import settings
from django.template import Context
from django.template.loader import get_template
from django.urls import reverse_lazy
from django.contrib.auth.views import PasswordResetView
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from rest_framework.permissions import IsAuthenticated
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
# from xhtml2pdf import pisa
# import pandas as pd
# from pandas import ExcelWriter
# from pandas import ExcelFile
# import numpy as np
# from openpyxl import load_workbook
# Create your views here.



# def link_callback(uri, rel):
#     """
#     Convert HTML URIs to absolute system paths so xhtml2pdf can access those
#     resources
#     """
#     # use short variable names
#     sUrl = settings.STATIC_URL      # Typically /static/
#     sRoot = settings.STATIC_ROOT    # Typically /home/userX/project_static/
#     mUrl = settings.MEDIA_URL       # Typically /static/media/
#     mRoot = settings.MEDIA_ROOT     # Typically /home/userX/project_static/media/

#     # convert URIs to absolute system paths
#     if uri.startswith(mUrl):
#         path = os.path.join(mRoot, uri.replace(mUrl, ""))
#     elif uri.startswith(sUrl):
#         path = os.path.join(sRoot, uri.replace(sUrl, ""))
#     else:
#         return uri  # handle absolute uri (ie: http://some.tld/foo.png)

#     # make sure that file exists
#     if not os.path.isfile(path):
#             raise Exception(
#                 'media URI must start with %s or %s' % (sUrl, mUrl)
#             )
#     return path




# def render_pdf_view(request,pk):
# 	enroll_stud=get_object_or_404(EnrollbyStudent,pk=pk)
# 	subjects=SubjectsLoaded.objects.filter(enrolled_by_student=enroll_stud,status="enrolled")
# 	template_path = 'student_profile/user_printer.html'
# 	now = datetime.now(timezone.utc)
# 	total_units=0
# 	for sub in subjects:
# 		total_units=total_units+sub.subject.unit
# 	context = {'enroll': enroll_stud,'subjects': subjects,'time':now.strftime('%b %d, %Y'),'total_units':total_units}
# 	# Create a Django response object, and specify content_type as pdf
# 	response = HttpResponse(content_type='application/pdf')
# 	# response['Content-Disposition'] = 'attachment; filename="report.pdf"'
# 	# find the template and render it.
# 	template = get_template(template_path)
# 	html = template.render(context)

# 	# create a pdf
# 	pisaStatus = pisa.CreatePDF(
# 	html, dest=response, link_callback=link_callback)
# 	# if error then show some funy view
# 	if pisaStatus.err:
# 		return HttpResponse('We had some errors <pre>' + html + '</pre>')
# 	return response



from django.shortcuts import render
from .models import Student
from django.contrib.auth.decorators import login_required

@login_required
def profile(request):
    # Attempt to fetch the student profile for the logged-in user
    student_profile = Student.objects.filter(user=request.user).first()  # Returns None if no profile exists

    # Pass the student profile (or None) to the template
    return render(request, 'student_profile/student_profile.html', {'student': student_profile})




@login_required(login_url=settings.LOGIN_URL)
def update_profile(request):
    try:
        student_profile = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        student_profile = None

    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student_profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('view_profile')  
        else:
            messages.error(request, 'Error updating your profile. Please check the form for errors.')
    else:
        form = StudentForm(instance=student_profile)

    context = {
        'form': form,
    }
    return render(request, 'student_profile/profile_update.html', context)






def subject_leaderboard(request):
    return render(request, 'student_profile/subject_leaderboard.html')

from rest_framework import generics
from .models import SubjectsLoaded
from .serializers import SubjectsLoadedSerializer

from django.db.models import Min

from rest_framework import generics
from .models import SubjectsLoaded, GradeSubmissionSchedule
from .serializers import SubjectsLoadedSerializers


class ForwardedSubmitStatusView(generics.ListAPIView):
    serializer_class = SubjectsLoadedSerializers

    def get_queryset(self):
        # Fetch the active semester and academic year
        active_schedule = GradeSubmissionSchedule.objects.first()

        if active_schedule:
            active_semister = active_schedule.active_semister
            active_academic_year = active_schedule.academic_year

            # Fetch subjects that match the active semester, year, and have any non-empty submit_status
            all_entries = SubjectsLoaded.objects.filter(
                semister=active_semister,
                academic_year=active_academic_year
            ).exclude(submit_status__isnull=True).exclude(submit_status__exact='')  # Exclude null/empty statuses

            # Filter subjects based on IDandFullname relationships
            filtered_entries = []
            for entry in all_entries:
                if IDandFullname.objects.filter(
                    instructor=entry.instructor,
                    subjects=entry.subject
                ).exists():
                    filtered_entries.append(entry)

            # Remove duplicates
            seen = set()
            unique_entries = []
            for entry in filtered_entries:
                key = (entry.subject_id, entry.instructor_id)
                if key not in seen:
                    seen.add(key)
                    unique_entries.append(entry)

            return unique_entries

        # Return no entries if there's no active semester or academic year
        return SubjectsLoaded.objects.none()












from django.shortcuts import render, redirect
from .forms import GradeSubmissionScheduleForm
from .models import GradeSubmissionSchedule, ActiveSem, Academic_year

def set_grade_submission_schedule(request):
    schedule, created = GradeSubmissionSchedule.objects.get_or_create(pk=1)

    if request.method == 'POST':
        form = GradeSubmissionScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.update_status()
            schedule.save()
            return redirect('set_grade_submission_schedule')

    form = GradeSubmissionScheduleForm(instance=schedule)
    context = {
        'form': form,
        'schedule': schedule,
        'active_semisters': ActiveSem.objects.all(),
        'academic_years': Academic_year.objects.all(),
    }
    return render(request, 'student_profile/set_grade_submission_schedule.html', context)



# for adding user (regestrar side)....................
# @login_required
# def add_role(request, role):
#     courses = Course.objects.all()  # Get all courses
    
#     if request.method == 'POST':
#         form = IDsNameForm(request.POST)
#         if form.is_valid():
#             id_fullname = form.save(commit=False)  # Create instance but don't save yet
#             id_fullname.added_by = request.user  # Set the user who added the role

#             # Set the role based on the view
#             if role == 'teacher':
#                 id_fullname.is_teacher = True
#             elif role == 'dean':
#                 id_fullname.is_dean = True
#             elif role == 'student':
#                 id_fullname.is_student = True

#             # Correct the logic for saving the selected course
#             course_id = request.POST.get('course_id')
#             if course_id:
#                 try:
#                     course = Course.objects.get(id=course_id)  # Fetch the Course object using the selected ID
#                     id_fullname.course = course
#                 except Course.DoesNotExist:
#                     pass

#             id_fullname.save()  # Save the instance to the database

#             # Automatically associate the instructor if the role is teacher or dean
#             if id_fullname.is_teacher or id_fullname.is_dean:
#                 instructor = Instructor.objects.create(first_name=id_fullname.first_name, last_name=id_fullname.last_name)
#                 id_fullname.instructor = instructor
#                 id_fullname.save()  # Save again to update with the instructor info

#             messages.success(request, f"{role.capitalize()} added successfully!")
#             return redirect('add_role', role=role)  # Redirect after post

#     else:
#         form = IDsNameForm()  # Initialize an empty form

#     return render(request, 'student_profile/add_role.html', {
#         'form': form,
#         'role': role,
#         'role_entries': IDandFullname.objects.filter(
#             added_by=request.user,
#             **{f'is_{role}': True}
#         ),
#         'courses': courses  # Pass courses to the template
#     })



import random

from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import IDsNameForm
from .models import Course, IDandFullname, Instructor

def generate_random_id():
    """Generate a random ID in the format 123-456 (digits only)."""
    return f"{random.randint(1000, 9999)}"

@login_required
def add_role(request, role):
    courses = Course.objects.all()  # Get all courses
    
    # Generate random ID only for teacher and dean roles
    random_id = None
    if role in ['teacher', 'dean']:
        random_id = generate_random_id()

    if request.method == 'POST':
        form = IDsNameForm(request.POST)
        if form.is_valid():
            id_fullname = form.save(commit=False)  # Create instance but don't save yet
            id_fullname.added_by = request.user  # Set the user who added the role

            # Set the role based on the view
            if role == 'teacher':
                id_fullname.is_teacher = True
            elif role == 'dean':
                id_fullname.is_dean = True
            elif role == 'student':
                id_fullname.is_student = True

            # Correct the logic for saving the selected course
            course_id = request.POST.get('course_id')
            if course_id:
                try:
                    course = Course.objects.get(id=course_id)  # Fetch the Course object using the selected ID
                    id_fullname.course = course
                except Course.DoesNotExist:
                    pass

            id_fullname.save()  # Save the instance to the database

            # Automatically associate the instructor if the role is teacher or dean
            # if id_fullname.is_teacher or id_fullname.is_dean:
            #     instructor = Instructor.objects.create(first_name=id_fullname.first_name, last_name=id_fullname.last_name)
            #     id_fullname.instructor = instructor
            #     id_fullname.save()  # Save again to update with the instructor info

            messages.success(request, f"{role.capitalize()} added successfully!")
            return redirect('add_role', role=role)  # Redirect after post

    else:
        form = IDsNameForm()  # Initialize an empty form

    return render(request, 'student_profile/add_role.html', {
        'form': form,
        'role': role,
        'role_entries': IDandFullname.objects.filter(
            added_by=request.user,
            **{f'is_{role}': True}
        ),
        'courses': courses,  # Pass courses to the template
        'random_id': random_id  # Pass the generated ID to the template, only if it's teacher or dean
    })



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import IDandFullname

@csrf_exempt
def delete_entry(request):
    if request.method == "POST":
        entry_id = request.POST.get("id")
        try:
            entry = IDandFullname.objects.get(id=entry_id)
            entry.delete()
            return JsonResponse({"success": True, "message": "Entry deleted successfully!"})
        except IDandFullname.DoesNotExist:
            return JsonResponse({"success": False, "message": "Entry not found."})
    return JsonResponse({"success": False, "message": "Invalid request."})


from django.shortcuts import render
from .models import IDandFullname

@login_required
def view_role_details(request, entry_id):
    # Fetch the entry based on ID
    entry = IDandFullname.objects.get(id=entry_id)
    return render(request, 'student_profile/add_role.html', {'entry': entry})



from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import IDandFullname

def delete_role_entry(request, entry_id):
    """Delete a role entry and return success response."""
    if request.method == 'POST':
        entry = get_object_or_404(IDandFullname, id=entry_id)
        entry.delete()  # Delete the entry
        return JsonResponse({'status': 'success', 'message': 'Entry deleted successfully'})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)




from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

@login_required
def delete_role(request, role, user_id):
    """
    Delete a user based on their ID and role.
    """
    try:
        # Filter user based on role and ID
        user = get_object_or_404(IDandFullname, id=user_id, **{f'is_{role}': True})
        user.delete()  # Delete the user
        return JsonResponse({'success': True, 'message': f'{role.capitalize()} deleted successfully!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})



@login_required
def profiles(request):
    # Fetch the student profile associated with the logged-in use
    # student_profile = get_object_or_404(Student, user=request.user)
    instructor_profile = Instructor.objects.filter(user=request.user).first()  # Returns None if no profile exists

    # Pass the student profile to the template
    return render(request, 'student_profile/instructor_profile.html', {'instructor': instructor_profile})



from django.shortcuts import render
from .models import IDandFullname, Course
import json

def teacher_list(request):
    # Get the teacher list including the course code
    teacher_list = IDandFullname.objects.filter(added_by=request.user).select_related('course').values(
        'id', 'student_id', 'first_name', 'last_name', 'course__code', 'added_by__username'  # Fetch course code directly
    )
    teachers_json = json.dumps(list(teacher_list))  # Convert to a JSON string

    return render(request, 'student_profile/teacher_list.html', {
        'teachers': teachers_json  # Pass teachers to the template
    })

def manage_subjects(request, teacher_id):
    teacher = get_object_or_404(IDandFullname, id=teacher_id)
    assigned_subjects = teacher.subjects.all()  # Get all subjects assigned to the teacher
    return render(request, 'student_profile/manage_subject.html', {'teacher': teacher, 'assigned_subjects': assigned_subjects})




from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import IDandFullname, Subject

# Fetch all subjects for a teacher
def get_teacher_subjects(request, teacher_id):
    teacher = get_object_or_404(IDandFullname, id=teacher_id)
    subjects = teacher.subjects.values('id', 'code', 'description', 'unit')
    return JsonResponse(list(subjects), safe=False)

# Search for subjects by code or description
def search_subjects(request):
    query = request.GET.get('query', '')
    subjects = Subject.objects.filter(code__icontains=query) | Subject.objects.filter(description__icontains=query)
    subjects = subjects.values('id', 'code', 'description', 'unit')[:10]
    return JsonResponse(list(subjects), safe=False)
    
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import IDandFullname, Subject

# Delete a subject assigned to a teacher
def delete_teacher_subject(request, teacher_id, subject_id):
    teacher = get_object_or_404(IDandFullname, id=teacher_id)
    subject = get_object_or_404(Subject, id=subject_id)
    
    # Check if the subject is assigned to the teacher
    if subject in teacher.subjects.all():
        teacher.subjects.remove(subject)
        teacher.save()
        return JsonResponse({'success': True, 'message': f'Subject {subject.code} removed successfully.'})
    else:
        return JsonResponse({'success': False, 'message': 'Subject not assigned to this teacher.'})


def teacher_subject(request, student_id):
    teacher = get_object_or_404(IDandFullname, student_id=student_id)
    return render(request, 'student_profile/teacher_subject.html', {'teacher': teacher})





from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import IDandFullname, Subject
from django.views.decorators.csrf import csrf_exempt
import json

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from .models import IDandFullname, Subject
import json

@csrf_exempt
def select_subject(request, teacher_id):
    teacher = get_object_or_404(IDandFullname, id=teacher_id)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            subject_code = data.get('subjectCode')

            if not subject_code:
                return JsonResponse({'success': False, 'message': 'Subject code is required.'})

            subject = get_object_or_404(Subject, code=subject_code)

            if teacher.subjects.filter(id=subject.id).exists():
                return JsonResponse({'success': False, 'message': f'{teacher.first_name} {teacher.last_name} is already assigned to {subject.description}.'})

            teacher.subjects.add(subject)

            return JsonResponse({
                'success': True,
                'message': f'Subject {subject.code} added successfully!',
                'id': subject.id,
                'name': subject.description
            })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data.'})

    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


def manage_subject(request, teacher_id):
    teacher = get_object_or_404(IDandFullname, id=teacher_id)
    assigned_subjects = teacher.subjects.all()  # Get all subjects assigned to the teacher
    return render(request, 'student_profile/manage_subject.html', {'teacher': teacher, 'assigned_subjects': assigned_subjects})



def get_teacher_subjects(request, teacher_id):
    teacher = get_object_or_404(IDandFullname, id=teacher_id)
    subjects = teacher.subjects.all()

    subject_list = [
        {
            'id': subject.id,
            'name': subject.description,  
            'subject': subject.code,
            'status': 'Assigned'
        } for subject in subjects
    ]

    return JsonResponse(subject_list, safe=False)






@api_view(['GET'])
def subject_list(request):
    search_query = request.GET.get('search', '')
    subjects = Subject.objects.filter(
        code__icontains=search_query
    ) | Subject.objects.filter(
        description__icontains=search_query
    )
    
    serializer = SubjectSerializer(subjects, many=True)
    return Response(serializer.data)




# def teacher_subject(request, teacher_id):
#     teacher = get_object_or_404(IDandFullname, id=teacher_id)
#     filtered_teachers = IDandFullname.objects.filter(is_teacher=True)

#     if request.method == 'POST':
#         action = request.POST.get('action')

#         if action == 'search_subjects':
#             query = request.POST.get('query', '')
#             if query:
#                 subjects = Subject.objects.filter(
#                     Q(code__icontains=query) | Q(description__icontains=query)
#                 )[:5]
#                 subject_list = [{
#                     'code': subject.code,
#                     'description': subject.description,
#                     'credits': subject.unit
#                 } for subject in subjects]
#                 return JsonResponse(subject_list, safe=False)

#         elif action == 'search_user_subjects':
#             query = request.POST.get('query', '')
#             if query:
#                 user_subjects = teacher.subjects.filter(
#                     Q(code__icontains=query) | Q(description__icontains=query)
#                 )[:5]
#                 user_subject_list = [{
#                     'id': subject.id,
#                     'code': subject.code,
#                     'description': subject.description
#                 } for subject in user_subjects]
#                 return JsonResponse(user_subject_list, safe=False)

#     related_subjects = teacher.subjects.all()

#     return render(request, 'student_profile/teacher_subject.html', {
#         'teacher': teacher,
#         'related_subjects': related_subjects,
#         'filtered_teachers': filtered_teachers,
#     })










# def teacher_subject(request, teacher_id):
#     teacher = get_object_or_404(IDandFullname, id=teacher_id)
#     filtered_teachers = IDandFullname.objects.filter(is_teacher=True)

#     if request.method == 'POST':
#         action = request.POST.get('action')

#         if action == 'search_subjects':
#             query = request.POST.get('query', '')
#             if query:
#                 subjects = Subject.objects.filter(
#                     Q(code__icontains=query) | Q(description__icontains=query)
#                 )[:5]
#                 subject_list = [{
#                     'code': subject.code,
#                     'description': subject.description,
#                     'credits': subject.unit
#                 } for subject in subjects]
#                 return JsonResponse(subject_list, safe=False)

#         elif action == 'search_user_subjects':
#             query = request.POST.get('query', '')
#             if query:
#                 user_subjects = teacher.subjects.filter(
#                     Q(code__icontains=query) | Q(description__icontains=query)
#                 )[:5]
#                 user_subject_list = [{
#                     'id': subject.id,
#                     'code': subject.code,
#                     'description': subject.description
#                 } for subject in user_subjects]
#                 return JsonResponse(user_subject_list, safe=False)

#         # Existing actions (add, edit, delete subjects)...
#         elif action == 'add_subject':
#             subject_code = request.POST.get('subject_code')
#             subject_description = request.POST.get('subject_description')
#             credits = request.POST.get('credits')

#             # Validate inputs
#             if not subject_code or not subject_description or not credits:
#                 return JsonResponse({'success': False, 'message': 'Please provide all required information: subject code, description, and credits.'})

#             try:
#                 # Fetch the subject using the provided data
#                 subject = Subject.objects.get(
#                     code=subject_code, 
#                     description=subject_description, 
#                     unit=credits
#                 )

#                 # Check if the subject is already assigned to the current teacher
#                 if teacher.subjects.filter(id=subject.id).exists():
#                     return JsonResponse({'success': False, 'message': 'This subject is already assigned to you.'})

#                 # Associate the subject with the teacher
#                 teacher.subjects.add(subject)

#                 return JsonResponse({'success': True, 'message': 'Subject added successfully!'})

#             except Subject.DoesNotExist:
#                 return JsonResponse({'success': False, 'message': 'Subject not found.'})

#         elif action == 'edit_subject':
#             subject_id = request.POST.get('subject_id')
#             subject_code = request.POST.get('subject_code')
#             subject_description = request.POST.get('subject_description')
#             credits = request.POST.get('credits')

#             try:
#                 subject = Subject.objects.get(id=subject_id)
#                 subject.code = subject_code
#                 subject.description = subject_description
#                 subject.unit = credits
#                 subject.save()
#                 return JsonResponse({'success': True, 'message': 'Subject updated successfully!'})
#             except Subject.DoesNotExist:
#                 return JsonResponse({'success': False, 'message': 'Subject not found.'})

#         elif action == 'delete_subject':
#             subject_id = request.POST.get('subject_id')
#             try:
#                 subject = Subject.objects.get(id=subject_id)
#                 teacher.subjects.remove(subject)  # Remove the subject association
#                 return JsonResponse({'success': True, 'message': 'Subject removed successfully!'})
#             except Subject.DoesNotExist:
#                 return JsonResponse({'success': False, 'message': 'Subject not found.'})

#     related_subjects = teacher.subjects.all()

#     return render(request, 'student_profile/teacher_subject.html', {
#         'teacher': teacher,
#         'related_subjects': related_subjects,
#         'filtered_teachers': filtered_teachers,
#     })
    







def manageGradesubmition(request):
	return render(request,'student_profile/manage_gradesubmition.html',{})

    
def scheduleGradesubmition(request):
	return render(request,'student_profile/subgrade_schedule.html',{})

def teacherSubject(request):
	return render(request,'student_profile/teacherSubject.html',{})



def templateDownload(request):
	return render(request,'student_profile/template_download.html',{})



# dean side.........
# from django.http import JsonResponse
# from accounts.models import CustomUser

# from django.http import JsonResponse
# from .models import CustomUser

# def get_teachers(request):
#     # Fetch users where is_teacher is True, include student_id
#     teachers = CustomUser.objects.filter(is_teacher=True).values('id', 'student_id', 'first_name', 'last_name')
#     teacher_list = list(teachers)  # Convert queryset to list
#     return JsonResponse(teacher_list, safe=False)

# from django.shortcuts import render, get_object_or_404
# from django.http import JsonResponse, HttpResponseBadRequest
# from .models import CustomUser
# from django.views.decorators.csrf import csrf_exempt
# import json

# # Get details of a teacher for editing
# def edit_teacher(request, id):
#     if request.method == 'GET':
#         # Get the teacher by id
#         teacher = get_object_or_404(CustomUser, id=id, is_teacher=True)
#         teacher_data = {
#             'student_id': teacher.student_id,
#             'first_name': teacher.first_name,
#             'last_name': teacher.last_name,
#             'email': teacher.email,
#             # Add more fields if necessary
#         }
#         return JsonResponse(teacher_data)

#     # Update teacher data
#     elif request.method == 'POST':
#         try:
#             teacher = get_object_or_404(CustomUser, id=id, is_teacher=True)
#             data = json.loads(request.body)
            
#             # Update teacher fields
#             teacher.first_name = data.get('first_name', teacher.first_name)
#             teacher.last_name = data.get('last_name', teacher.last_name)
#             teacher.email = data.get('email', teacher.email)
#             teacher.save()

#             return JsonResponse({'message': 'Teacher details updated successfully.'})
#         except Exception as e:
#             return HttpResponseBadRequest(f"Error: {str(e)}")








import os
import json
from django.core.exceptions import ValidationError
from django.core.files.storage import default_storage
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, get_object_or_404
import pandas as pd
import logging
from django.conf import settings
from .models import Student, SubjectsLoaded, EnrollbyStudent
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter  
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from django.http import HttpResponse


# Configure logging
logger = logging.getLogger(__name__)

# for uploading grades teacher side.......


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Instructor

# def enrollstud(request,pk):
# 	return render(request,'student_profile/enrollstud.html',{'student_id':pk})

from .models import GradeSubmissionSchedule

@csrf_exempt
def upload_and_displaypage(request, student_id):
    student_profile = get_object_or_404(IDandFullname, student_id=student_id)
    instructor = student_profile.instructor
    
    # Get the status of the first GradeSubmissionSchedule record
    schedule = GradeSubmissionSchedule.objects.first()
    status = schedule.status if schedule else 'pending'  # Default to 'pending' if no records found

    context = {
        'instructor_name': f"{instructor.first_name} {instructor.last_name}" if instructor else "Instructor not assigned",
        'instructor_id': instructor.id if instructor else None,
        'status': status
    }
    return render(request, 'student_profile/upload_grades.html', context)
import os
import logging
import pandas as pd
from django.core.exceptions import ValidationError
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.conf import settings
import chardet 

logger = logging.getLogger(__name__)

def get_file_encoding(file_path):
    """Automatically detect the encoding of the file using chardet."""
    with open(file_path, 'rb') as f:
        raw_data = f.read()
    result = chardet.detect(raw_data)
    return result['encoding']

def handle_uploaded_file(file):
    try:
        if not file:
            raise ValidationError("No file provided.")
        
        file_path = os.path.join(settings.MEDIA_ROOT, file.name)
        logger.info(f"Attempting to save file: {file.name} to {file_path}")
        
        with default_storage.open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        file_encoding = get_file_encoding(file_path)
        logger.info(f"Detected file encoding: {file_encoding}")

        if file.name.endswith('.csv'):
            logger.info("Reading CSV file")
            try:
                df = pd.read_csv(file_path, encoding=file_encoding, on_bad_lines='skip')  
            except pd.errors.ParserError as e:
                logger.error(f"Error reading CSV file: {e}")
                raise ValidationError("Error reading CSV file.")
        elif file.name.endswith('.xlsx'):
            logger.info("Reading Excel file, second sheet")
            try:
                df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl') 
            except Exception as e:
                logger.error(f"Error reading Excel file: {e}")
                raise ValidationError("Error reading Excel file.")
        else:
            logger.error("Unsupported file format.")
            raise ValidationError("Unsupported file format. Only .csv and .xlsx are allowed.")

        if df.empty:
            logger.error("The file is empty.")
            raise ValidationError("The uploaded file is empty.")
        
        logger.info(f"File loaded into DataFrame, first few rows:\n{df.head()}")
        df.columns = df.columns.str.lower()

        required_columns = ['student_id', 'name', 'midterm_rating', 'final_rating', 'status']
        missing_columns = [col for col in required_columns if col.lower() not in df.columns]

        if missing_columns:
            logger.error(f"Missing columns: {', '.join(missing_columns)}")  
            raise ValidationError(f"Missing columns: {', '.join(missing_columns)}")  

        logger.info("File successfully validated and processed.")
        return df[required_columns], file_path

    except Exception as e:
        logger.error(f"Error in handle_uploaded_file: {e}")
        raise e

@csrf_exempt
def upload_and_display(request):
    if request.method == 'POST':
        try:
            file = request.FILES.get('file')
            if not file:
                logger.error("No file uploaded")
                return JsonResponse({'error_message': 'No file uploaded.'}, status=400)

            logger.info(f"Received file: {file.name}")
            df, file_path = handle_uploaded_file(file)

            if df.empty:
                logger.warning("Uploaded file is empty")
                return JsonResponse({'error_message': 'No data found in the file.'}, status=400)

            student_ids = df['student_id'].unique()

            students_in_db = set(Student.objects.filter(student_id__in=student_ids).values_list('student_id', flat=True))

            missing_students = set(df['student_id']) - students_in_db

            if missing_students:
                error_message = f"Missing students: {', '.join(missing_students)}."
                logger.warning(f"Missing records: {error_message}")
                return JsonResponse({'error_message': error_message}, status=400)

            table_html = df.to_html(classes='table table-bordered table-striped table-hover', index=False)
            return JsonResponse({
                'table_html': table_html,
                'error_message': None,
                'data': df.to_json(orient='records'),
                'file_path': file_path
            })

        except ValidationError as e:
            logger.error(f"Validation error: {e}")
            return JsonResponse({'error_message': str(e)}, status=400)
        except Exception as e:
            logger.error(f"Unexpected error in upload_and_display: {e}")
            return JsonResponse({'error_message': f"Error processing file: {str(e)}"}, status=500)


from django.db import transaction
from datetime import datetime
from django.http import JsonResponse

@csrf_exempt
def save_grades(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            grades = data.get('data')
            selected_subject_id = data.get('selectedSubject')
            instructor_id = data.get('instructorId')  

            if not grades:
                return JsonResponse({'error_message': 'No grades data provided.'}, status=400)

            grades = json.loads(grades)
            selected_subject = get_object_or_404(Subject, id=selected_subject_id)
            instructor = get_object_or_404(Instructor, id=instructor_id)

            active_academic_year = Academic_year.objects.filter(status=True).first()
            if not active_academic_year:
                return JsonResponse({'error_message': 'No active academic year found.'}, status=400)

            active_semister = ActiveSem.objects.first()
            if not active_semister:
                return JsonResponse({'error_message': 'No active semister found.'}, status=400)

            instructor_record = IDandFullname.objects.filter(instructor=instructor, subjects=selected_subject).first()

            if not instructor_record:
                return JsonResponse({'error_message': f'Instructor {instructor.first_name} {instructor.last_name} is not assigned to subject {selected_subject.code}.'}, status=400)

            existing_grades = SubjectsLoaded.objects.filter(subject=selected_subject, instructor=instructor, submit_status='forwarded').exists()

            if existing_grades:
                return JsonResponse({'error_message': 'Grades for this subject have already been submitted by this instructor.'}, status=400)


            with transaction.atomic():
                for entry in grades:
                    student_id = entry['student_id']
                    midterm_grade = entry['midterm_grade']
                    grade = entry['final_grade']
                    status = entry['status']

                    student = get_object_or_404(Student, student_id=student_id)
                    enrollments = EnrollbyStudent.objects.filter(student=student)

                    if not enrollments.exists():
                        return JsonResponse({'error_message': f'No enrollment found for student {student_id}.'}, status=400)

                    enrolled_by_student = enrollments.first()

                    SubjectsLoaded.objects.update_or_create(
                        enrolled_by_student=enrolled_by_student,
                        subject=selected_subject,
                        defaults={
                            'instructor': instructor,
                            'midterm_grade': midterm_grade,
                            'grade': grade,
                            'grade_status': status,
                            'submit_status': 'forwarded',
                            'submission_time': datetime.now(),
                            'academic_year': active_academic_year,
                            'semister': active_semister  
                        }
                    )

            return JsonResponse({
                'message': 'Grades saved successfully and marked as forwarded.',
                'academic_year': active_academic_year.ay,
                'semister': active_semister.semister,
                'submission_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })

        except Exception as e:
            logger.error(f"Error saving grades: {e}")
            return JsonResponse({'error_message': 'Error saving grades. Please check your data.'}, status=400)

    return JsonResponse({'error_message': 'Invalid request.'}, status=400)



from django.db.models import Q


class SubjectViewSet(viewsets.ViewSet):
    
    def list(self, request):
        search_query = request.GET.get('search', '')

        subjects = Subject.objects.filter(
            Q(code__icontains=search_query) | Q(description__icontains=search_query)
        )[:10]
        
        serializer = SubjectSerializer(subjects, many=True)
        return Response(serializer.data)





@csrf_exempt
def check_existing_grades(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            grades = data.get('grades')
            selected_subject_id = data.get('selectedSubject')  
            existing_records = []

            # Ensure selected subject exists
            selected_subject = get_object_or_404(Subject, id=selected_subject_id)

            for entry in grades:
                student_id = entry['student_id']

                # Check if the student has any non-null grades for the selected subject
                if SubjectsLoaded.objects.filter(
                    enrolled_by_student__student__student_id=student_id, 
                    subject=selected_subject
                ).exclude(midterm_grade__isnull=True, grade__isnull=True).exclude(midterm_grade='', grade='').exists():
                    existing_records.append({
                        'student_id': student_id,
                        'subject': selected_subject.code,  # Use selected subject code
                        'name': entry.get('name', 'Unknown')  # Adjust as needed
                    })

            return JsonResponse({'existing_records': existing_records})

        except Exception as e:
            logger.error(f"Error checking existing grades: {e}")
            return JsonResponse({'error_message': 'Error checking existing grades.'}, status=400)

    return JsonResponse({'error_message': 'Invalid request.'}, status=400)



# for downloading template..............


from rest_framework import viewsets
from rest_framework.response import Response
from .models import SubjectsLoaded
from .serializers import StudentSerializer 


from django.db.models import Q

class EnrolledStudentsViewSet(viewsets.ViewSet):
    def retrieve(self, request, pk=None):
        # Filter students enrolled in the subject and exclude those who already have a grade
        students = SubjectsLoaded.objects.filter(
            subject_id=pk,
            grade__isnull=True  # Assuming `grade` field exists and null means no grade
        ).select_related('enrolled_by_student__student')

        enrolled_students = [
            {
                'student_id': subject.enrolled_by_student.student.student_id,  # Fetching real student ID
                'name': f"{subject.enrolled_by_student.student.last_name}, {subject.enrolled_by_student.student.first_name} ",
            }
            for subject in students
        ]

        return Response(enrolled_students)


import openpyxl 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl.styles import Protection

class GenerateExcelAPIView(APIView):
    def post(self, request):
        students = request.data.get('students', [])
        subject_name = request.data.get('subject_name', 'Enrolled Students')
        

        if not students:
            return Response({"error": "No students selected"}, status=400)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'[:31]

        logo_path = 'student_profile/templates/student_profile/bcc.png'
        logo = Image(logo_path)
        logo.width, logo.height = 80, 80
        ws.add_image(logo, 'D1')

        ws.merge_cells('C1:N1')
        ws['C1'] = "BUENAVISTA COMMUNITY COLLEGE"
        ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].font = Font(bold=True, name='Times New Roman', size=14)

        ws.merge_cells('C2:N2')
        ws['C2'] = '"Caring Your Future"'
        ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C2'].font = Font(bold=True, italic=True, name='Times New Roman', size=12)

        ws.merge_cells('C3:N3')
        ws['C3'] = "Cangawa, Buenavista, Bohol"
        ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C3'].font = Font(name='Times New Roman', size=12)

        ws.merge_cells('C5:N5')
        ws['C5'] = "COLLEGE DEPARTMENT RATING SHEET"
        ws['C5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C5'].font = Font(bold=True, name='Times New Roman', size=12)

        ws.merge_cells('C6:N6')
        ws['C6'] = "  Course:____________________________________     Year: _______________     Term & A.Y.: ____________"
        ws['C6'].alignment = Alignment(horizontal='left', vertical='center')
        ws['C6'].font = Font(bold=True, name='Times New Roman', size=10)
        ws.row_dimensions[6].height = 25

        ws.merge_cells('C7:N7')
        ws['C7'] = f' Subject Code: ______ Descriptive Title: __{subject_name}__   Units: ______'
        ws['C7'].alignment = Alignment(horizontal='left', vertical='center')
        ws['C7'].font = Font(bold=True, name='Times New Roman', size=10)
        ws.row_dimensions[7].height = 25

        headers = [
            'Student ID', 'Name of Student', '   ', '   ', 'Prelim', 'Midterm', 'Average', ' ', 'Semi Final', 'Final', 'Average', ' ', 'Final Rating', 'Status'
        ]

        gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        for col_num, header in enumerate(headers, start=2):
            if header == 'Student ID':
                ws.cell(row=8, column=2, value=header).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row=8, column=2).font = Font(bold=True, name='Times New Roman', size=9)
                ws.cell(row=8, column=2).border = Border(
                    top=Side(border_style='thin', color='000000'),
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                ws.cell(row=8, column=2).fill = gray_fill

            elif header == 'Name of Student':
                ws.merge_cells('C8:E8')
                ws.cell(row=8, column=3, value="Name of Student").alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.cell(row=8, column=3).font = Font(bold=True, name='Times New Roman', size=9)
                ws.cell(row=8, column=3).border = Border(
                    top=Side(border_style='thin', color='000000'),
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                ws.cell(row=8, column=3).fill = gray_fill
                
                ws.cell(row=8, column=4).border = Border(
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                ws.cell(row=8, column=4).fill = gray_fill

                ws.cell(row=8, column=5).border = Border(
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                ws.cell(row=8, column=5).fill = gray_fill

            else:
                if col_num >= 6:
                    ws.cell(row=8, column=col_num, value=header).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.cell(row=8, column=col_num).font = Font(bold=True, name='Times New Roman', size=9)
                    ws.cell(row=8, column=col_num).border = Border(
                        top=Side(border_style='thin', color='000000'),
                        left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000')
                    )
                    ws.cell(row=8, column=col_num).fill = gray_fill

        ws.row_dimensions[8].height = 25

        one_decimal_style = NamedStyle(name="one_decimal")
        one_decimal_style.number_format = '0.0'

        TOTAL_ROWS = 20

        for row_idx, student in enumerate(students, start=9):
            ws.cell(row=row_idx, column=2, value=student['student_id'])
            ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=5)
            ws.cell(row=row_idx, column=3, value=student['name'])

            for col_idx in range(6, 12):
                ws.cell(row=row_idx, column=col_idx, value='')

            average_h = ws.cell(row=row_idx, column=8, value=f"=IF(AND(ISNUMBER(F{row_idx}), ISNUMBER(G{row_idx})), ROUND(AVERAGE(F{row_idx}:G{row_idx}),1), \"\")")
            average_h.style = one_decimal_style

            average_l = ws.cell(row=row_idx, column=12, value=f"=IF(AND(ISNUMBER(J{row_idx}), ISNUMBER(K{row_idx})), ROUND(AVERAGE(J{row_idx}:K{row_idx}),1), \"\")")
            average_l.style = one_decimal_style

            final_rating = ws.cell(row=row_idx, column=14, value=f"=IF(AND(ISNUMBER(H{row_idx}), ISNUMBER(L{row_idx})), ROUND(AVERAGE(H{row_idx},L{row_idx}),1), \"\")")
            final_rating.style = one_decimal_style

            status = ws.cell(row=row_idx, column=15, value=f'=IF(ISNUMBER(N{row_idx}), IF(N{row_idx}<3.1,"Passed","Failed"), "Incomplete")')

            for col_idx in range(2, 15):
                ws.cell(row=row_idx, column=col_idx).border = Border(
                    top=Side(border_style='thin', color='000000'),
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
        end_row = len(students) + 9
        ws.cell(row=end_row, column=2, value="END")
        ws.merge_cells(start_row=end_row, start_column=3, end_row=end_row, end_column=5)
        ws.cell(row=end_row, column=3, value="")


        for col_idx in range(2, 15):
            ws.cell(row=end_row, column=col_idx).border = Border(
                top=Side(border_style='thin', color='000000'),
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )

        blank_rows_needed = TOTAL_ROWS - (len(students) + 1) 
        if blank_rows_needed > 0:
            for i in range(blank_rows_needed):
                blank_row = end_row + 1 + i
                ws.cell(row=blank_row, column=2, value="") 
                ws.merge_cells(start_row=blank_row, start_column=3, end_row=blank_row, end_column=5)  # Merge Name of Student
                ws.cell(row=blank_row, column=3, value="") 

             
                for col_idx in range(2, 15):
                    ws.cell(row=blank_row, column=col_idx).border = Border(
                        top=Side(border_style='thin', color='000000'),
                        left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000')
                    )


     
        if "one_decimal" not in wb.named_styles:
            wb.add_named_style(one_decimal_style)

        column_widths = [10, 7, 12, 11, 8, 8, 8, 1, 8, 8, 8, 1, 8, 10]  
        for i, width in enumerate(column_widths, start=2):  
            ws.column_dimensions[get_column_letter(i)].width = width




        ws.print_area = 'C1:N{0}'.format(len(students) + 18 + blank_rows_needed) 
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT 
        ws.page_setup.fitToWidth = 1  
        ws.page_setup.fitToHeight = 1  


        footer_text = "Certified Correct:"
        footer_row = len(students) + 11 + blank_rows_needed
        ws.merge_cells(start_row=footer_row, start_column=3, end_row=footer_row, end_column=14)
        ws.cell(row=footer_row, column=3, value=footer_text)
        ws.cell(row=footer_row, column=3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=footer_row, column=3).font = Font(name='Times New Roman', size=10, italic=True, bold=True)
        ws.row_dimensions[footer_row].height = 25  


        footer_text2 = "  _____________________                                                                              _____________________"
        footer_row2 = footer_row + 1
        ws.merge_cells(start_row=footer_row2, start_column=3, end_row=footer_row2, end_column=14)
        ws.cell(row=footer_row2, column=3, value=footer_text2)
        ws.cell(row=footer_row2, column=3).alignment = Alignment(horizontal='center', vertical='bottom')
        ws.cell(row=footer_row2, column=3).font = Font(name='Times New Roman', size=10, italic=True)
        ws.row_dimensions[footer_row2].height = 25  


        footer_text3 = "   Instructor                                                                                               date submitted"
        footer_row3 = footer_row2 + 1
        ws.merge_cells(start_row=footer_row3, start_column=3, end_row=footer_row3, end_column=14)
        ws.cell(row=footer_row3, column=3, value=footer_text3)
        ws.cell(row=footer_row3, column=3).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=footer_row3, column=3).font = Font(name='Times New Roman', size=10, italic=True)
        ws.row_dimensions[footer_row3].height = 25

        footer_text4 = "Noted by:                                                                                            Aprroved by:    DR. ANALIZA L. CHUA  "
        footer_row4 = footer_row3 + 1
        ws.merge_cells(start_row=footer_row4, start_column=3, end_row=footer_row4, end_column=14)
        ws.cell(row=footer_row4, column=3, value=footer_text4)
        ws.cell(row=footer_row4, column=3).alignment = Alignment(horizontal='center', vertical='bottom')
        ws.cell(row=footer_row4, column=3).font = Font(name='Times New Roman', size=10, italic=True, bold=True)
        ws.row_dimensions[footer_row4].height = 25 


        footer_text5 = "              Department Dean                                                                                         College President"
        footer_row5 = footer_row4 + 1 
        ws.merge_cells(start_row=footer_row5, start_column=3, end_row=footer_row5, end_column=14)
        ws.cell(row=footer_row5, column=3, value=footer_text5)
        ws.cell(row=footer_row5, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=footer_row5, column=3).font = Font(name='Times New Roman', size=10, italic=True)
        ws.row_dimensions[footer_row5].height = 15


        footer_text6 = "Recorded by:           MA. MAY N. CUPTA            "
        footer_row6 = footer_row5 + 2
        ws.merge_cells(start_row=footer_row6, start_column=3, end_row=footer_row6, end_column=14)
        ws.cell(row=footer_row6, column=3, value=footer_text6)
        ws.cell(row=footer_row6, column=3).alignment = Alignment(horizontal='center', vertical='bottom')
        ws.cell(row=footer_row6, column=3).font = Font(name='Times New Roman', size=10, italic=True, bold=True)
        ws.row_dimensions[footer_row6].height = 29

        footer_text7 = "                           College Registrar I             "
        footer_row7 = footer_row6 + 1
        ws.merge_cells(start_row=footer_row7, start_column=3, end_row=footer_row7, end_column=14)
        ws.cell(row=footer_row7, column=3, value=footer_text7)
        ws.cell(row=footer_row7, column=3).alignment = Alignment(horizontal='center', vertical='bottom')
        ws.cell(row=footer_row7, column=3).font = Font(name='Times New Roman', size=10, italic=True)
        ws.row_dimensions[footer_row7].height = 12
        



        ws2 = wb.create_sheet(title="Sheet2")
        ws2.protection.sheet = True
        ws2.protection.enable() 
        wb._sheets = [wb["Sheet2"]] + [sheet for sheet in wb._sheets if sheet.title != "Sheet2"]

        wb.active = 1
        wb["Sheet2"].sheet_state = "hidden"

        

        # Copy headers and the structure from Sheet1 to Sheet2
        # Copy headers, data, and styles from Sheet1 to Sheet2
        for row_idx in range(8, len(students) + 8 + 1):  # Start from row 8 to cover headers
            for row_idx in range(8, len(students) + 8 + 1):  # Start from row 8 to cover headers
                for col_idx in range(2, 15):  # From column 2 (Student ID) to column 15 (Status)
                    # Copy data from Sheet1 to Sheet2
                    ws2.cell(row=row_idx, column=col_idx).value = f"=Sheet1!{get_column_letter(col_idx)}{row_idx}"

                    # Copy style from Sheet1 to Sheet2
                    cell_style = ws[get_column_letter(col_idx) + str(row_idx)]._style
                    ws2[get_column_letter(col_idx) + str(row_idx)]._style = cell_style

            # Hide columns 4 and 5
            ws2.column_dimensions[get_column_letter(4)].hidden = True
            ws2.column_dimensions[get_column_letter(5)].hidden = True

            ws2[get_column_letter(2) + '8'].value = 'student_id'
            ws2[get_column_letter(3) + '8'].value = 'name'
            ws2[get_column_letter(8) + '8'].value = 'Midterm_Rating'
            ws2[get_column_letter(14) + '8'].value = 'Final_Rating'
            ws2[get_column_letter(15) + '8'].value = 'status'


        # Now link the grades between Sheet1 and Sheet2, keeping the same style for the grade columns
        for row_idx, student in enumerate(students, start=9):
            # Link Prelim (Column 6)
            ws2.cell(row=row_idx, column=6).value = f"=Sheet1!F{row_idx}"
            ws2[get_column_letter(6) + str(row_idx)]._style = ws[get_column_letter(6) + str(row_idx)]._style
            
            # Link Midterm (Column 7)
            ws2.cell(row=row_idx, column=7).value = f"=Sheet1!G{row_idx}"
            ws2[get_column_letter(7) + str(row_idx)]._style = ws[get_column_letter(7) + str(row_idx)]._style
            
            # Link Semi Final (Column 9)
            ws2.cell(row=row_idx, column=9).value = f"=Sheet1!J{row_idx}"
            ws2[get_column_letter(9) + str(row_idx)]._style = ws[get_column_letter(9) + str(row_idx)]._style
            
            # Link Final (Column 10)
            ws2.cell(row=row_idx, column=10).value = f"=Sheet1!K{row_idx}"
            ws2[get_column_letter(10) + str(row_idx)]._style = ws[get_column_letter(10) + str(row_idx)]._style
            
            # Link Average (Column 8 for Prelim-Midterm, Column 12 for Semi-Final-Final)
            ws2.cell(row=row_idx, column=8).value = f"=Sheet1!H{row_idx}"
            ws2[get_column_letter(8) + str(row_idx)]._style = ws[get_column_letter(8) + str(row_idx)]._style

            ws2.cell(row=row_idx, column=12).value = f"=Sheet1!L{row_idx}"
            ws2[get_column_letter(12) + str(row_idx)]._style = ws[get_column_letter(12) + str(row_idx)]._style
            
            # Link Final Rating (Column 14)
            ws2.cell(row=row_idx, column=14).value = f"=Sheet1!N{row_idx}"
            ws2[get_column_letter(14) + str(row_idx)]._style = ws[get_column_letter(14) + str(row_idx)]._style

            # Link Status (Column 15)
            ws2.cell(row=row_idx, column=15).value = f"=Sheet1!O{row_idx}"
            ws2[get_column_letter(15) + str(row_idx)]._style = ws[get_column_letter(15) + str(row_idx)]._style

        # Adjust column widths in Sheet2
        for i, width in enumerate(column_widths, start=2):  
            ws2.column_dimensions[get_column_letter(i)].width = width

        # Set the header row height for Sheet2 for consistency
        ws2.row_dimensions[8].height = 25




        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{subject_name}_grades.csv"'
        wb.save(response)
        return response

















def terms(request):
    return render(request, 'student_profile/terms.html')
def privacy(request):
    return render(request, 'student_profile/privacy.html')



# for checking pre rquisite................
from django.http import JsonResponse
from .models import Subject, SubjectsLoaded

def check_prerequisites(request):
    subject_id = request.GET.get('subject_id')
    student_id = request.GET.get('student_id')
    
    try:
        subject = Subject.objects.get(id=subject_id)
    except Subject.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Subject does not exist'})

    prerequisite = subject.prerequisite  

    if not prerequisite:
        return JsonResponse({'status': 'success', 'message': 'No prerequisites for this subject'})
    
    has_grade = SubjectsLoaded.objects.filter(
        enrolled_by_student__student__id=student_id,
        subject=prerequisite
    ).exclude(
        grade__isnull=True
    ).exclude(
        grade__exact=''
    ).exists()

    if not has_grade:
        return JsonResponse({'status': 'error', 'message': f'Prerequisite {prerequisite.code} not completed'})
    
    return JsonResponse({'status': 'success', 'message': 'Prerequisite satisfied'})



class enrollFormView(AjaxFormMixin, FormView):
    form_class = EnrollForm
    template_name  = 'forms/ajax.html'
    success_url = '/form-success/'


@login_required
def post_announcement(request):
    if request.user.is_superuser: 
        if request.method == 'POST':
            form = AnnouncementForm(request.POST)
            if form.is_valid():
                announcement = form.save(commit=False)
                announcement.created_by = request.user
                announcement.save()
                return redirect('announcement_list')
        else:
            form = AnnouncementForm()
        return render(request, 'student_profile/post_announcement.html', {'form': form})
    else:
        return redirect('announcement_list')

def announcement_list(request):
    now = timezone.now()
    if request.user.is_superuser:
        announcements = Announcement.objects.all().order_by('-start_date')
    else:
        announcements = Announcement.objects.filter(
            start_date__lte=now,
            end_date__gte=now
        ).order_by('-start_date')
    return render(request, 'student_profile/announcement_list.html', {'announcements': announcements})

@require_POST
def update_announcement(request):
    form = AnnouncementForm(request.POST)
    if form.is_valid():
        announcement_id = request.POST.get('id')
        announcement = get_object_or_404(Announcement, id=announcement_id)
        announcement.title = form.cleaned_data['title']
        announcement.content = form.cleaned_data['content']
        announcement.start_date = form.cleaned_data['start_date']
        announcement.end_date = form.cleaned_data['end_date']
        announcement.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Invalid data'})

@login_required
@require_POST
def delete_announcement(request):
    if request.user.is_superuser:
        announcement_id = request.POST.get('id')
        announcement = get_object_or_404(Announcement, id=announcement_id)
        announcement.delete()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'You do not have permission to delete this announcement.'})

class AnnouncementViewSet(viewsets.ModelViewSet):
    serializer_class = AnnouncementSerializer
    permission_classes = [IsAuthenticated]
    queryset = Announcement.objects.all()  

    def get_queryset(self):
        now = timezone.now()
        user = self.request.user
        if user.is_superuser:
            return Announcement.objects.all().order_by('-start_date')
        else:
            return Announcement.objects.filter(
                start_date__lte=now,
                end_date__gte=now
            ).order_by('-start_date')




class enrollFormView(AjaxFormMixin, FormView):
    form_class = EnrollForm
    template_name  = 'forms/ajax.html'
    success_url = '/form-success/'


@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def enrollSTudAPI(request,pk):

	if request.method=='POST':
		form=EnrollForm(request.POST)
		if form.is_valid():
			student=get_object_or_404(Student,pk=pk)
			enroll_key=student.student_id+str(form.cleaned_data['academic_year'])+str(form.cleaned_data['course'])+str(form.cleaned_data['semister'])
			student_enrollKey=EnrollbyStudent.objects.filter(enrollment_key=enroll_key)
			if not student_enrollKey.exists():
				enroll=form.save(commit=False)
				enroll.student=student
				enroll.enrollment_key=enroll_key
				enroll.save()


				enrollRes={'pk': enroll.pk,'message':'enrolled successfully','is_save':True}
				return Response(enrollRes)
			else:
				enrollRes={'message':'already enrolled on this semister','is_save':False}
				return Response(enrollRes)
		return Response({'message':'invalid forms'})
	return Response({'message':'get request'})



@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def updateStatusAPI(request,pk):
	enroll=get_object_or_404(EnrollbyStudent,pk=pk)
	subjects=SubjectsLoaded.objects.filter(enrolled_by_student=enroll)
	if request.method=='POST':
		if subjects.exists():
			enroll.status='approved'
			enroll.date_created=datetime.now(timezone.utc)
			enroll.save()
			serialize_subjects=[]
			for subject_loaded in subjects:
				subject_loaded.status="enrolled"
				subject_loaded.save()


				subject_enroll={'pk':str(subject_loaded.pk),'subject_pk':str(subject_loaded.subject.pk),'code':subject_loaded.subject.code,
					'description':subject_loaded.subject.description,'unit':subject_loaded.subject.unit,'status':subject_loaded.status}
				serialize_subjects.append(subject_enroll)

			return Response({'message':'ok','enroll_status':'approved','subjects':serialize_subjects})
		return Response({'message':'you don\'t have subjects added'})
	return Response({'message':'get request'})



@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def returnForCorrectionAPI(request,pk):
	enroll=get_object_or_404(EnrollbyStudent,pk=pk)
	subjects=SubjectsLoaded.objects.filter(enrolled_by_student=enroll)
	if request.method=='POST':
		if subjects.exists():
			enroll.status='returned for correction'
			enroll.save()
			serialize_subjects=[]
			for subject_loaded in subjects:
				subject_loaded.status="returned for correction"
				subject_loaded.save()


				subject_enroll={'pk':str(subject_loaded.pk),'subject_pk':str(subject_loaded.subject.pk),'code':subject_loaded.subject.code,
					'description':subject_loaded.subject.description,'unit':subject_loaded.subject.unit,'status':subject_loaded.status}
				serialize_subjects.append(subject_enroll)

			return Response({'message':'ok','enroll_status':enroll.status,'subjects':serialize_subjects})
		return Response({'message':'you don\'t have subjects added'})
	return Response({'message':'get request'})




@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def displayAllEnrollAPI(request):
	enrollees=EnrollbyStudent.objects.all()
	enroll_dic_list=[]

	for enrollee in enrollees:
		enroll_dic_list.append({'student_id':str(enrollee.student.student_id),
			'fullname': enrollee.student.first_name + " " + enrollee.student.middle_name +" "+enrollee.student.last_name,
			'course':enrollee.course.code, 'ay':enrollee.academic_year.ay,'semister':enrollee.semister,'status':enrollee.status})
	context={'enrollees':enroll_dic_list}
	return Response(context)







@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def forwardForApprovalAPI(request,pk):
	enroll=get_object_or_404(EnrollbyStudent,pk=pk)
	subjects=SubjectsLoaded.objects.filter(enrolled_by_student=enroll)
	if request.method=='POST':
		if subjects.exists():
			enroll.status='forwarded'
			enroll.save()
			serialize_subjects=[]
			for subject_loaded in subjects:
				subject_loaded.status="forwarded"
				subject_loaded.save()


				subject_enroll={'pk':str(subject_loaded.pk),'subject_pk':str(subject_loaded.subject.pk),'code':subject_loaded.subject.code,
					'description':subject_loaded.subject.description,'unit':subject_loaded.subject.unit,'status':subject_loaded.status}
				serialize_subjects.append(subject_enroll)

			return Response({'is_forwarded': True,'message':'forward successfully','enroll_status':'forwarded','subjects':serialize_subjects})
		return Response({'message':'you don\'t have subjects added','is_forwarded': False})
	return Response({'message':'get request'})







def viewEnrolledStudentAdmin(request):
	enroll_students=EnrollbyStudent.objects.select_related("student").select_related("year_level").select_related("academic_year").select_related("major").select_related("course").all()

	context={'enroll_students':enroll_students}
	

	return render(request,'student_profile/enroll_students.html',{})


def manageGrade(request):
	return render(request,'student_profile/manage_grades.html',{})

def about_us(request):
	return render(request,'student_profile/about_us.html',{})



	
def viewEnrolledStudent(request,pk):
	student=get_object_or_404(Student,pk=pk)
	enroll_students=EnrollbyStudent.objects.filter(student=student)
	context={'enroll_students':enroll_students,'student':student}
	return render(request,'student_profile/byStudentEnrollList.html',context)
def viewEnrolledStudentByStudentId(request,student_id):
	student=get_object_or_404(Student,student_id=student_id)
	enroll_students=EnrollbyStudent.objects.filter(student=student)
	context={'enroll_students':enroll_students,'student':student}
	return render(request,'student_profile/byStudentEnrollList.html',context)

def enrollstud(request,pk):
	return render(request,'student_profile/enrollstud.html',{'student_id':pk})

@api_view(['GET'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def getEnrolledCount(request, pk):
	student = get_object_or_404(Student,pk=pk)
	enroll_by_students = EnrollbyStudent.objects.filter(student=student, status= 'approved')
	return Response({'enroll_count':len(enroll_by_students)})






class ByCourseStudentEnrollList(generics.ListAPIView):
	serializer_class = StudentsEnrollSerializer
	permission_classes = (IsAuthenticated,)
	authentication_classes=(SessionAuthentication,)

	def get_queryset(self):
		ay_pk=self.kwargs['ay']
		course_pk=self.kwargs['course']
		sem=self.kwargs['sem']
		major_pk = self.kwargs['major']
		if major_pk != "0":
			course=get_object_or_404(Course,pk=course_pk)
			ay=get_object_or_404(Academic_year,pk=ay_pk)
			major = get_object_or_404(Major,pk=major_pk)
			return EnrollbyStudent.objects.select_related("student").select_related("year_level").select_related("academic_year").select_related("major").select_related("course").filter(semister=sem,course=course,academic_year=ay,status="approved", major=major).order_by('year_level','major', 'student__last_name')
		else:
			course=get_object_or_404(Course,pk=course_pk)
			ay=get_object_or_404(Academic_year,pk=ay_pk)
	
			return EnrollbyStudent.objects.select_related("student").select_related("year_level").select_related("academic_year").select_related("major").select_related("course").filter(semister=sem,course=course,academic_year=ay,status="approved").order_by('year_level','major', 'student__last_name')
	


class SubjectsList(generics.ListAPIView):
	serializer_class = SubjectSerializer
	permission_classes = (IsAuthenticated,)
	authentication_classes=(SessionAuthentication,)

	def get_queryset(self):
		return Subject.objects.all()

class StudentEnrollListByAcademicYear(generics.ListAPIView):
	serializer_class = StudentEnrollSerializer
	
	def get_queryset(self):
		ayId = self.kwargs['ay']
		ay = get_object_or_404(Academic_year,pk=ayId)
		return EnrollbyStudent.objects.select_related("student").select_related("year_level").select_related("academic_year").select_related("major").select_related("course").filter(academic_year=ay)

class StudentEnrollListByStudentId(generics.ListAPIView):
	serializer_class = StudentEnrollSerializer
	
	def get_queryset(self):

		student = get_object_or_404(Student,pk=self.kwargs['pk'])
		return EnrollbyStudent.objects.select_related("student").select_related("year_level").select_related("academic_year").select_related("major").select_related("course").filter(student=student)


class StudentEnrollList(generics.ListAPIView):
	serializer_class = StudentEnrollSerializer
	
	def get_queryset(self):
		"""
		This view should return a list of all the purchases
		for the currently authenticated user.
		"""
		return EnrollbyStudent.objects.select_related("student").select_related("year_level").select_related("academic_year").select_related("major").select_related("course").all()

class AllStudentEnrollBySYAndSEM(generics.ListAPIView):
	serializer_class = StudentEnrollSerializer
	
	def get_queryset(self):
		sem=self.kwargs['sem']
		ayId = self.kwargs['ay']
		coursePk = self.kwargs['course']
		yearPk = self.kwargs['year']
		majorPk = self.kwargs['major']
		ay = get_object_or_404(Academic_year,pk=ayId)
		course = get_object_or_404(Course,pk=coursePk)
		year = get_object_or_404(Year_level,pk=yearPk)
		results = []
		if majorPk != '-1':
			major = get_object_or_404(Major, pk = majorPk)
			results = EnrollbyStudent.objects.select_related("student").select_related("year_level").select_related("academic_year").select_related("major").select_related("course").filter(major = major,academic_year=ay, semister = sem, course = course, year_level= year, status ='approved').all()
			if len(results) != 0:
				return EnrollbyStudent.objects.select_related("student").select_related("year_level").select_related("academic_year").select_related("major").select_related("course").filter(major = major,academic_year=ay, semister = sem, course = course, year_level= year, status ='approved').all()
			
			return EnrollbyStudent.objects.select_related("student").select_related("year_level").select_related("academic_year").select_related("major").select_related("course").filter(academic_year=ay, semister = sem, course = course, year_level= year, status ='approved').all()
				
		return EnrollbyStudent.objects.select_related("student").select_related("year_level").select_related("academic_year").select_related("major").select_related("course").filter(academic_year=ay, semister = sem, course = course, year_level= year, status ='approved').all()
		

class SubjectbyCurriculum(generics.ListAPIView):
	serializer_class= SubjectSerializer
	def get_queryset(self):
		major_pk=self.kwargs['major']
		course_pk=self.kwargs['course']
		sem=self.kwargs['sem']
		yearpk=self.kwargs['year']
		if major_pk != "0":
			year=get_object_or_404(Year_level,pk=yearpk)
			major=get_object_or_404(Major,pk=major_pk)
			course=get_object_or_404(Course,pk=course_pk)
			return  Curriculum.objects.filter(course=course,year_level=year,semister=sem,major=major).first().subjects.all()

		
		year=get_object_or_404(Year_level,pk=yearpk)
		course=get_object_or_404(Course,pk=course_pk)
		return Curriculum.objects.filter(course=course,year_level=year,semister=sem).first().subjects.all()


class SearchEnroll(generics.ListAPIView):
	serializer_class= EnrollStudSerlizer
	def get_queryset(self):
		major_pk=self.kwargs['major']
		course_pk=self.kwargs['course']
		sem=self.kwargs['sem']
		year=self.kwargs['year']
		ay = self.kwargs['ay']
		student=get_object_or_404(Student,pk=self.kwargs['id'])

		if major_pk != "0":
			
			major=get_object_or_404(Major,pk=major_pk)
			course=get_object_or_404(Course,pk=course_pk)
			year=get_object_or_404(Year_level,pk=year)
			academic_year = get_object_or_404(Academic_year,pk = ay)
			return EnrollbyStudent.objects.filter(academic_year = academic_year, student=student,semister=sem,course=course,year_level=year,major=major)


		
		course=get_object_or_404(Course,pk=course_pk)
		year=get_object_or_404(Year_level,pk=year)
		academic_year = get_object_or_404(Academic_year,pk = ay)
		return EnrollbyStudent.objects.filter(academic_year = academic_year,student=student,semister=sem,course=course,year_level=year)




class SubjectLoadedList(generics.ListAPIView):
	serializer_class= SubjectEnrollSerializer
	def get_queryset(self):
		sub_pk=self.kwargs['sub']
		major_pk=self.kwargs['major']
		course_pk=self.kwargs['course']
		sem=self.kwargs['sem']
		ay=self.kwargs['ay']

		if major_pk != "0":
			sub=get_object_or_404(Subject,pk=sub_pk)
			major=get_object_or_404(Major,pk=major_pk)
			course=get_object_or_404(Course,pk=course_pk)
			academic_y=get_object_or_404(Academic_year,pk=ay)
			return SubjectsLoaded.objects.select_related("enrolled_by_student").filter(enrolled_by_student__course=course,enrolled_by_student__academic_year=academic_y,enrolled_by_student__semister=sem,subject=sub,enrolled_by_student__major=major,status="enrolled")
		if major_pk == "0" and course_pk == "0":
			sub=get_object_or_404(Subject,pk=sub_pk)
			academic_y=get_object_or_404(Academic_year,pk=ay)
			return SubjectsLoaded.objects.select_related("enrolled_by_student").filter(enrolled_by_student__academic_year=academic_y,enrolled_by_student__semister=sem,subject=sub,status="enrolled")
	
		
		sub=get_object_or_404(Subject,pk=sub_pk)
		course=get_object_or_404(Course,pk=course_pk)
		academic_y=get_object_or_404(Academic_year,pk=ay)
		return SubjectsLoaded.objects.select_related("enrolled_by_student").filter(enrolled_by_student__course=course,enrolled_by_student__academic_year=academic_y,enrolled_by_student__semister=sem,subject=sub,status="enrolled")

class SubjectByInstructorList(generics.ListAPIView):
	serializer_class= SubjectEnrollSerializer
	def get_queryset(self):
		instruc=self.kwargs['instructor']
		sem=self.kwargs['sem']
		ay=self.kwargs['ay']

      
		instructor=get_object_or_404(Instructor,pk=instruc)
		academic_y=get_object_or_404(Academic_year,pk=ay)
		return SubjectsLoaded.objects.select_related("enrolled_by_student").filter(enrolled_by_student__academic_year=academic_y,enrolled_by_student__semister=sem,instructor=instructor,status="enrolled")

class SubjectByInstructorListAdmin(generics.ListAPIView):
    serializer_class = SubjectEnrollSerializer

    def get_queryset(self):
        instructor_id = self.kwargs['instructor']
        subject_id = self.kwargs.get('subject')  # Use subject ID from URL

        instructor = get_object_or_404(Instructor, pk=instructor_id)
        queryset = SubjectsLoaded.objects.select_related("enrolled_by_student").filter(
            instructor=instructor,
            subject__pk=subject_id,  # Filter by subject ID
            status="enrolled"
        )
        return queryset
        
def gradesDetailsAdmin(request, instructor_pk, subject_pk):
    serializer_class = SubjectEnrollSerializer
    form = SubjectLoadForm()
    enroll = get_object_or_404(Instructor, pk=instructor_pk)
    subjects = SubjectsLoaded.objects.filter(instructor=enroll, subject__pk=subject_pk)

    context = {'enroll': enroll, 'form': form, 'subjects': subjects}
    return render(request, 'student_profile/manage_gradesubmitionAdmin.html', context)



class SubjectByStudentList(generics.ListAPIView):
	serializer_class= SubjectEnrollSerializer
	def get_queryset(self):
		stud=self.kwargs['student']
		sem=self.kwargs['sem']
		ay=self.kwargs['ay']

		student=get_object_or_404(Student,pk=stud)
		academic_y=get_object_or_404(Academic_year,pk=ay)
		return SubjectsLoaded.objects.select_related("enrolled_by_student").filter(enrolled_by_student__student=student,enrolled_by_student__academic_year=academic_y,enrolled_by_student__semister=sem,status="enrolled")


class IntructorLoadedSubjectList(generics.ListAPIView):
	serializer_class= InstructorLoadSubjectWithObjectSerializer
	def get_queryset(self):
		instruc=self.kwargs['instructor']
		sem=self.kwargs['sem']
		ay=self.kwargs['ay']

		instructor=get_object_or_404(Instructor,pk=instruc)
		academic_y=get_object_or_404(Academic_year,pk=ay)
		return InstructorLoadSubject.objects.filter(instructor=instructor,academic_year=academic_y,semister=sem)

class IntructorLoadedSubjectByClassList(generics.ListAPIView):
	serializer_class= InstructorLoadSubjectWithObjectSerializer
	def get_queryset(self):
		course_pk=self.kwargs['course']
		sem=self.kwargs['sem']
		ay=self.kwargs['ay']
		section = self.kwargs['section']
		major_pk = self.kwargs['major']
		year_level_pk = self.kwargs['year_level']
		year_level = get_object_or_404(Year_level,pk=year_level_pk)
		course=get_object_or_404(Course,pk=course_pk)
		academic_y=get_object_or_404(Academic_year,pk=ay)
		if major_pk == '-1':
			return InstructorLoadSubject.objects.filter(course=course,academic_year=academic_y,semister=sem, section=section, year_level = year_level)
		major = get_object_or_404(Major,pk=major_pk)
		return InstructorLoadSubject.objects.filter(course=course,academic_year=academic_y,semister=sem, section=section, year_level = year_level, major=major)



class GetStudentIdAnfFullnameList(generics.ListAPIView):
	serializer_class= IDandFullnameSerializer
	def get_queryset(self):
		student_id=self.kwargs['student_id']
		return IDandFullname.objects.filter(student_id=student_id)




		



class EnrollStudViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
	
    queryset = EnrollbyStudent.objects.all()
    serializer_class = EnrollStudSerlizer

class IdAndFullnameViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = IDandFullname.objects.all()
    serializer_class = IDandFullnameSerializer


class CurriculumViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = Curriculum.objects.all()
    serializer_class = CurriculumSerializer

class CurriculumsViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = Curriculum.objects.all()
    serializer_class = CurriculumsSerializer



class CourseViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = Course.objects.all()
    serializer_class = CourseSerializer

class StudentViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = Student.objects.select_related("course").select_related("major").all()
    serializer_class = StudentSerializer

class MajorViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = Major.objects.all()
    serializer_class = MajorSerializer

class AyViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = Academic_year.objects.all()
    serializer_class = Academic_yearSerializer


class InstructorViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = Instructor.objects.all()
    serializer_class = InstructorSerializer

class SubjectEnrollViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = SubjectsLoaded.objects.all()
    serializer_class = SubjectsLoadedSerializer

class YearLevelViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = Year_level.objects.all()
    serializer_class = Year_levelSerializer

class AOSFViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = AOSF.objects.all()
    serializer_class = AOSFSerializer

class ActiveSemViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = ActiveSem.objects.all()
    serializer_class = ActiveSemSerializer

class SubjectViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer

class RoomViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = Room.objects.all()
    serializer_class =RoomSerializer

class DaySchedViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = DaySched.objects.all()
    serializer_class =DaySchedSerializer

class TimeSchedViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = TimeSched.objects.all()
    serializer_class =TimeSchedSerializer

class InstructorLoadSubjectViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = InstructorLoadSubject.objects.all()
    serializer_class = InstructorLoadSubjectSerializer

class InstructorLoadSubjectWithObjectViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows messages to be viewed or edited.
    """
    queryset = InstructorLoadSubject.objects.all()
    serializer_class = InstructorLoadSubjectWithObjectSerializer

@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def addSubjectLoadAPI(request,pk):

	enrollbyStud=get_object_or_404(EnrollbyStudent,pk=pk)
	enroll_subject=SubjectsLoaded.objects.filter(enrolled_by_student=enrollbyStud)


	if request.method == 'POST':
		form=SubjectLoadForm(request.POST)

		if form.is_valid():

			addSubject=form.save(commit=False)
			duplicate_subject=enroll_subject.filter(subject=addSubject.subject)
			if not duplicate_subject.exists():
				addSubject.enrolled_by_student=enrollbyStud
				addSubject.status="draft"
				addSubject.save()
				
				context={
						'pk':str(addSubject.pk),
						'subject_pk':str(addSubject.subject.pk),
						'code':str(addSubject.subject.code),
						'description':str(addSubject.subject.description),
						'unit':str(addSubject.subject.unit),
						'status':str(addSubject.status),
						'is_save':True}
				return Response(context)
			return Response({'message':'subject already enroll','is_save':False})
		return Response({'message':'fail','is_save':False})
	return Response({'message':'get request'})



@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def updateSubjectLoadAPI(request):

	if request.method == 'POST':
		subject_pk=request.POST
		subject=get_object_or_404(SubjectsLoaded,pk=subject_pk['is_updated'])
		form = SubjectLoadForm(request.POST, instance=subject)
		if form.is_valid():
			addSubject=form.save(commit=False)
			addSubject.save()
	
			context={
						'pk':str(addSubject.pk),
						'subject_pk':str(addSubject.subject.pk),
						'code':str(addSubject.subject.code),
						'description':str(addSubject.subject.description),
						'unit':str(addSubject.subject.unit),
						'status':str(addSubject.status),
					'is_updated':True,}
			return Response(context)

		return Response({'is_updated':False})
	return Response({'message':'get request'})








@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def deleteSubjectLoaded(request,pk):
	subject=get_object_or_404(SubjectsLoaded,pk=pk)
	if request.method == 'POST':
		subject.delete()
		return Response({'is_delete':True})
	return Response({'is_delete':False})




@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def dropSubjectLoaded(request,pk):
	subject=get_object_or_404(SubjectsLoaded,pk=pk)
	if request.method == 'POST':
		subject.status="drop"
		subject.save()
		serializer=SubjectsLoadedSerializer(subject)
		return Response(serializer.data)
	return Response({'is_drop':False})


@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def EnrolldropSubjectLoaded(request,pk):
	subject=get_object_or_404(SubjectsLoaded,pk=pk)
	if request.method == 'POST':
		subject.status="enrolled"
		subject.save()
		serializer=SubjectsLoadedSerializer(subject)
		return Response(serializer.data)
	return Response({'is_enroll':False})



@login_required(login_url=settings.LOGIN_URL)
def home(request):
 
    forwarded = EnrollbyStudent.objects.filter(status='forwarded')
    grade_forwarded = SubjectsLoaded.objects.filter(submit_status='forwarded')


    unique_instructors = {}
    for grade in grade_forwarded:
        key = (grade.instructor.id, grade.subject.id) 
        if key not in unique_instructors:
            unique_instructors[key] = grade


    distinct_grade_forwarded = list(unique_instructors.values())

    return render(request, 'student_profile/index.html', {
        'user': request.user,
        'count': 0,
        'forwarded': forwarded,
        'grade_forwarded': distinct_grade_forwarded, 
    })


   

@login_required(login_url=settings.LOGIN_URL)
def addNewStudent(request):
    form = StudentForm()
    students = Student.objects.all()
    
    if request.method == "POST":
        form = StudentForm(request.POST or None)
        
        if form.is_valid():
            student = form.save(commit=False)  
            
          
            if request.user.is_authenticated:
                student.user = request.user  
            else:
                return redirect(settings.LOGIN_URL)

            student.save()  
            return redirect('enrollstud', student.pk)
    
    return render(request, 'student_profile/add_student.html', {'form': form, 'students': students})

@login_required(login_url=settings.LOGIN_URL)
def addNewStudentAPI(request):

	return render(request,'student_profile/add_student_api.html')

@login_required(login_url=settings.LOGIN_URL)
def addStudent(request):
	form=StudentForm()
	students=Student.objects.all()
	if request.method == "POST":

		form=StudentForm(request.POST or None)
		if form.is_valid():
			student=form.save(commit=False)
			student.save()
			return redirect('home')
	return render(request,'student_profile/addStudent.html',{'form':form,'students':students})

@login_required(login_url=settings.LOGIN_URL)
def addCuriculum(request):
    form=CurriculumForm()
    curriculums=Curriculum.objects.all()
    if request.method == 'POST':
        form=CurriculumForm(request.POST)
        if form.is_valid():
            curriculum=form.save()
            # updateUnique slug for query porposes
            if curriculum.major == None:
            	curriculum.slug=str(curriculum.course.pk)+str(curriculum.year_level.pk)+str(curriculum.semister)
            else:
            	curriculum.slug=str(curriculum.course.pk)+str(curriculum.year_level.pk)+str(curriculum.semister)+str(curriculum.major.pk)  
            curriculum.save()
            return redirect('addCuriculum')
    context={'form':form,'curriculums':curriculums}
    return render(request,'student_profile/addCurriculumAPI.html',context)

@login_required(login_url=settings.LOGIN_URL)
def editCurriculum(request,pk):
	curriculum=get_object_or_404(Curriculum,pk=pk)
	form=CurriculumForm(instance=curriculum)
	if request.method == 'POST':
		form=CurriculumForm(request.POST,instance=curriculum)
		if form.is_valid():
			cur=form.save(commit=False)
			if cur.major == None:
				cur.slug=str(cur.course.pk)+str(cur.year_level.pk)+str(cur.semister)
				cur.save()
			cur.save()

			return redirect('addCuriculum')

	context={'form':form}
	return render(request,'student_profile/editCurriculum.html',context)



def editStudent(request,pk):
	student=get_object_or_404(Student,pk=pk)
	form=StudentFormUpdate(instance=student)
	if request.method == 'POST':
		form = StudentFormUpdate(request.POST,instance=student)
		if form.is_valid():
			form.save()
			return redirect('allStudent')

	context={'form':form}
	return render(request,'student_profile/edit_student.html', context)


def editIDs(request,pk):
	stud_id=get_object_or_404(IDandFullname,pk=pk)
	form=IDsNameForm(instance=stud_id)
	if request.method == 'POST':
		form = IDsNameForm(request.POST,instance=stud_id)
		if form.is_valid():
			form.save()
			return redirect('allStudent')

	context={'form':form}
	return render(request,'student_profile/edit_IDs.html', context)


@login_required(login_url=settings.LOGIN_URL)
def addSubject(request):
	form_upload=UploadExcelForm()
	form_subject=SubjectForm()
	if request.method=='POST':
		form=UploadExcelForm(request.POST,request.FILES)
		if form.is_valid():
			excel=form.save(commit=False)
			excel.save()
			try:
				df = pd.read_excel(excel.file.path, sheet_name=0)
			except Exception as e:
				# Handle the error, maybe log it or send a message to the user
				print(f"Error reading Excel file: {e}")
				return redirect('addSubject')

			
			for cell in df.values:
				print('naka abot')
				subject_exists=Subject.objects.filter(code=cell[0],description=cell[1])
				if not subject_exists.exists():
					Subject.objects.create(code=cell[0],description=cell[1],unit=cell[2])



			return redirect('addSubject')
	context={'form_upload':form_upload,'form_subject':form_subject,'subjects':Subject.objects.all()}
	return render(request,'student_profile/addSubject.html',context)





# @login_required(login_url=settings.LOGIN_URL)
# def addIDandName(request):
# 	form_upload=UploadExcelForm()
# 	form_subject=SubjectForm()
# 	if request.method=='POST':
# 		form=UploadExcelForm(request.POST,request.FILES)
# 		if form.is_valid():
# 			excel=form.save(commit=False)
# 			excel.save()
# 			df = pd.read_excel('C:/Bitnami/djangostack-2.1.1-1/apps/django/django_projects/enrollment'+excel.file.url, sheet_name=0)  
			
# 			for cell in df.values:
# 				print('naka abot')
# 				id_exists=IDandFullname.objects.filter(student_id=cell[0],first_name=cell[1],last_name=cell[2])
# 				if not id_exists.exists():
# 					IDandFullname.objects.create(student_id=cell[0],first_name=cell[1],last_name=cell[2])



# 			return redirect('addSubject')
# 	context={'form_upload':form_upload,'form_subject':form_subject,'subjects':Subject.objects.all()}
# 	return render(request,'student_profile/addSubject.html',context)



@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def getSubjectsByCurriculumAPI(request,*args,**kwargs):
		
	queryset = Curriculum.objects.all()
	sem = kwargs['sem']
	course_pk=kwargs['course']
	course=Course.objects.get(pk=course_pk)


	curriculums=queryset.filter(course=course,semister=sem)

	subjects=[]
	for curriculum in curriculums:
		for subject in curriculum.subjects.all():
			serializer=SubjectSerializer(subject)
			subjects.append(serializer.data)



	context={'subjects':subjects}



	return Response(context)




@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def getSubjectsByCurriculumAPIAll(request,*args,**kwargs):


	# queryset = Curriculum.objects.all()
	# sem = kwargs['sem']
	# course_pk=kwargs['course']
	# course=Course.objects.get(pk=course_pk)


	curriculums=Curriculum.objects.prefetch_related('subjects').all()


	subjects=[]
	for curriculum in curriculums:
		for subject in curriculum.subjects.all():
			serializer=SubjectSerializer(subject)
			subjects.append(serializer.data)



	context={'subjects':subjects}



	return Response(context)




def StudentByCourse(request):


	# sy=request.GET.get('ay')
	# course_pk=request.GET.get('course')

	# ay=Academic_year.objects.get(pk=sy)
	# course=Course.objects.get(pk=course_pk)
	# semister=request.GET.get('semister')



	# subjects_loaded=SubjectsLoaded.objects.all()
	# student_enroll=[]

	# if request.method == "POST":
	# 	student_enrollExcel=[]
	# 	course=Course.objects.get(pk=request.POST.get('course_pk'))
	# 	ay=Academic_year.objects.get(pk=request.POST.get('ay_pk'))

	# 	students=EnrollbyStudent.objects.filter(course=course,status="approved",academic_year=ay,semister=request.POST.get('semister'))
	
	# 	for student in students:
	# 		subjects=SubjectsLoaded.objects.filter(enrolled_by_student=student,status="enrolled")
	# 		total_units=0
	# 		course_codes=""
	# 		course_descs=""
	# 		for subject in subjects:
				
	# 			#now.strftime('%b %d, %Y')
	# 			total_units=total_units+subject.subject.unit
	# 			if course_codes == "":
	# 				course_codes=course_codes+subject.subject.code
	# 			else:
	# 				course_codes=course_codes+","+subject.subject.code

	# 			if course_descs == "":
	# 				course_descs=course_descs+subject.subject.description

	# 			else:
	# 				course_descs=course_descs+","+subject.subject.description




				
	# 		if not student.major == None:	
	# 			course_str=str(student.course) +" "+str(student.major)
	# 		else:
	# 			course_str=str(student.course)


	# 		student_enrollExcel.append([student.student.student_id,student.student.last_name,
	# 								student.student.first_name,student.student.middle_name,
	# 								student.student.ext_name,student.student.gender,student.student.address,student.student.birth_date,
	# 								course_str,student.year_level,student.student.zip_code,
	# 								student.student.email_add,student.student.mobile_no,
	# 								total_units,course_codes,course_descs])

	# 	df=pd.DataFrame(student_enrollExcel)
	# 	print(os.environ['USERPROFILE'])
	# 	home = expanduser("C:/Users/yamamz/Desktop")
	# 	writer=ExcelWriter(home+"/Report/report.xlsx")
	# 	df.to_excel(writer,'sheet1',index=False,header=False)
	# 	writer.save()






	# students=EnrollbyStudent.objects.filter(course=course,status="approved",academic_year=ay,semister=semister)
	
	# for student in students:

	# 	subjects=SubjectsLoaded.objects.filter(enrolled_by_student=student,status="enrolled")
	# 	total_units=0
	# 	course_codes=""
	# 	course_descs=""
	# 	for subject in subjects:
			
	# 		#now.strftime('%b %d, %Y')
	# 		total_units=total_units+subject.subject.unit
	# 		if course_codes == "":
	# 			course_codes=course_codes+subject.subject.code
	# 		else:
	# 			course_codes=course_codes+","+subject.subject.code

	# 		if course_descs == "":
	# 			course_descs=course_descs+subject.subject.description

	# 		else:
	# 			course_descs=course_descs+","+subject.subject.description

	# 	student_enroll.append({'student_id':student.student.student_id,'first_name':student.student.first_name,
	# 							'last_name':student.student.last_name,'middle_name':student.student.middle_name,
	# 							'ext_name':student.student.ext_name,'gender':student.student.gender,'birth_date':student.student.birth_date,
	# 							'course':student.course,'year_level':student.year_level,'zip_code':student.student.zip_code,
	# 							'email_add':student.student.email_add,'mobile_no':student.student.mobile_no,
	# 							'total_units':total_units,'course_codes':course_codes,'course_descs':course_descs})

	# context={'enroll_students':student_enroll,'course':course,'semister':semister,'ay':ay}

	return render(request,'student_profile/byCourse.html',{})






def enrollSelection(request):
	ays=Academic_year.objects.all()
	courses=Course.objects.all()
	return render(request,'student_profile/byStudCattegory.html',{'ays':ays,'courses':courses})


	
def viewGrades(request,pk):
	enroll= get_object_or_404(EnrollbyStudent, pk = pk)
	return render(request,'student_profile/view_grades.html',{'enroll':enroll})
def viewGradesStudent(request,pk):
	enroll= get_object_or_404(EnrollbyStudent, pk = pk)
	return render(request,'student_profile/view_grade_student.html',{'enroll':enroll})




@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def Apiaddsubjectbycuriculum(request,*args,**kwargs):
	major_pk=kwargs['majorpk']

	try:
		course=Course.objects.get(pk=request.POST.get('coursepk'))


		year=Year_level.objects.get(pk=request.POST.get('yearpk'))
		sem=request.POST.get('semister')
		major=Major.objects.get(pk=major_pk)
		subjects_by_curriculum= Curriculum.objects.filter(course=course,year_level=year,semister=sem,major=major)
	except Major.DoesNotExist:
		major = None

	if major == None:
		subjects_by_curriculum= Curriculum.objects.filter(course=course,year_level=year,semister=sem)

	



		
		


	if subjects_by_curriculum.exists():
		subjects={'status':True,'subjects':subjects_by_curriculum[0].subjects.all(),'message':'subjects added'}
		return Response(subjects)
	return Response({'status':False,'message':'curriculum does not exist'})














@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def addsubjectbycuriculumAPI(request,pk):
	enrollbyStud=get_object_or_404(EnrollbyStudent,pk=pk)
	if request.method == "POST":
		major_pk=request.POST.get('majorpk')
		if major_pk == "":
			major_pk=0

		try:
			course=Course.objects.get(pk=request.POST.get('coursepk'))


			year=Year_level.objects.get(pk=request.POST.get('yearpk'))
			sem=request.POST.get('semister')
			major=Major.objects.get(pk=major_pk)
			subjects_by_curriculum= Curriculum.objects.filter(course=course,year_level=year,semister=sem,major=major)
		except Major.DoesNotExist:
			major = None

		if major == None:
			subjects_by_curriculum= Curriculum.objects.filter(course=course,year_level=year,semister=sem)



			
			


		if subjects_by_curriculum.exists():
			
			serialize_subjects=[]
			for subject in subjects_by_curriculum[0].subjects.all():
				duplicate_subject=SubjectsLoaded.objects.filter(enrolled_by_student=enrollbyStud,subject=subject)

				
				if not duplicate_subject.exists():
					subject_loaded=SubjectsLoaded.objects.create(enrolled_by_student=enrollbyStud,subject=subject,status="draft")
					subject_enroll={'pk':str(subject_loaded.pk),'subject_pk':str(subject_loaded.subject.pk),'code':subject_loaded.subject.code,
					'description':subject_loaded.subject.description,'unit':subject_loaded.subject.unit,'status':subject_loaded.status}
					serialize_subjects.append(subject_enroll)

			subjects={'status':True,'subjects':serialize_subjects,'message':'subjects added'}
			return Response(subjects)


					

			

		return Response({'status':False,'message':'curriculum does not exist'})

	return Response({'message':'get'})


@login_required(login_url=settings.LOGIN_URL)
def enrollChoices(request):
	return render(request,'student_profile/choices.html')






@api_view(['GET', 'POST'])
@authentication_classes((SessionAuthentication,))
@permission_classes((IsAuthenticated,))
def uploadSubjectsAPI(request):
	if request.method == 'POST':
		form=UploadExcelForm(request.POST,request.FILES)
		if form.is_valid():

			form.save()
			df = pd.read_excel('C:/Bitnami/djangostack-2.1.1-1/apps/django/django_projects/enrollment'+form.file.url, sheet_name=0)  
			for cell in df.values:
				Subject.objects.create(code=cell[0],description=cell[1],unit=cell[2])
				context={'is_save':True}
			return Response(context)

		return Response({'is_save':False})
	return Response({'message':'get'})



	







@login_required(login_url=settings.LOGIN_URL)
def enroll(request,pk):
	student=get_object_or_404(Student,pk=pk)
	form=EnrollForm()
	form2=SubjectLoadForm()
	context={'student':student,'form':form,'form2':form2}
	if request.method == 'POST':
		form= EnrollForm(request.POST)
		if form.is_valid():
			enroll=form.save(commit=False)
			enroll.student=student
			if not enroll.major == None:
				student_enroll=EnrollbyStudent.objects.filter(student=enroll.student,
															course=enroll.course,
															year_level=enroll.year_level,
															semister=enroll.semister,
															major=enroll.major)
			else:
				student_enroll=EnrollbyStudent.objects.filter(student=enroll.student,
															course=enroll.course,
															year_level=enroll.year_level,
															semister=enroll.semister)

			if not student_enroll.exists():	
				enroll.status='draft'
				enroll.save()
				return redirect('enrollDetails', pk=enroll.pk)


		student=get_object_or_404(Student,pk=pk)
		form=EnrollForm()

		context={'student':student,'form':form,'form2':form2,'message':'the student already enrolled in this course with the same acamemic year and semister'}
		return render(request,'student_profile/enroll.html',context)

	context={'student':student,'form':form}
	return render(request,'student_profile/enroll.html',context)


def enrollDetails(request,pk):
    form=SubjectLoadForm()

    enroll=get_object_or_404(EnrollbyStudent,pk=pk)
    subjects=SubjectsLoaded.objects.filter(enrolled_by_student=enroll)
    form2=EnrollForm(instance=enroll)

    if request.method == 'POST':
    	form2=EnrollForm(request.POST,instance=enroll)
    	if form2.is_valid():
    		form2.save()
    		return redirect('enrollDetails',enroll.pk)

    context={'enroll':enroll,'form':form,'form2':form2,'subjects':subjects}
    return render(request,'student_profile/enrolldetailss.html',context)


def enrollDetailsAdmin(request,pk):
    form=SubjectLoadForm()
    enroll=get_object_or_404(EnrollbyStudent,pk=pk)
    subjects=SubjectsLoaded.objects.filter(enrolled_by_student=enroll)

    # if request.method == 'POST':
    total_units=0
    for sub in subjects:
    	total_units=total_units+sub.subject.unit

    context={'enroll':enroll,'form':form,'subjects':subjects,'total_units':total_units}
    return render(request,'student_profile/enrollDetailsAdmins.html',context)


def checkMyEnrollment(request):
	if request.method == 'POST':
		student_id=request.POST.get('student_id')
		try:
			student=Student.objects.get(student_id=student_id)
		except Student.DoesNotExist:
			student=None

		if not student == None:
			return redirect('allStudentEnroll-student',student.pk)
		else:
			return render(request,'student_profile/check_my_enrollment.html',{'message':'Student ID doest not exists'})



	return render(request,'student_profile/check_my_enrollment.html',{})




def registerConfirm(request):
	return render(request,'student_profile/registerConfirmation.html')


def displayAllStudent(request):
	return render(request,'student_profile/allStudent.html',{})

def schedules(request):
	return render(request,'student_profile/schedules.html',{})
def users_page(request):
	return render(request,'student_profile/user_list.html',{})

def page_not_found(request,  exception):
	return render(request, "student_profile/404.html", {})


class ResetPasswordView(SuccessMessageMixin, PasswordResetView):
    template_name = 'student_profile/password_reset.html'
    email_template_name = 'student_profile/password_reset_email.html'
    subject_template_name = 'users/password_reset_subject'
    success_message = "We've emailed you instructions for setting your password, " \
                      "if an account exists with the email you entered. You should receive them shortly." \
                      " If you don't receive an email, " \
                      "please make sure you've entered the address you registered with, and check your spam folder."
    success_url = reverse_lazy('home')



class StudentListView(generics.ListAPIView):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['first_name', 'last_name', 'middle_name', 'student_id']

class UserListView(generics.ListAPIView):
	queryset = CustomUser.objects.all()
	serializer_class = UserSerializer
	filter_backends = [filters.SearchFilter]
	authentication_classes = [SessionAuthentication]
	permission_classes = [IsAuthenticated]
	search_fields = ['first_name', 'last_name', 'student_id']

class UserViewSet(viewsets.ModelViewSet):
	authentication_classes = [SessionAuthentication]
	permission_classes = [IsAuthenticated]
	queryset = CustomUser.objects.all()
	serializer_class = UserSerializer



def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect('change_password')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'accounts/change_password.html', {
        'form': form
    })

